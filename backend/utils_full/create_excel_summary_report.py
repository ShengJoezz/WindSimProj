# @Author: joe 847304926@qq.com
# @Date: 2025-07-06 16:17:25
# @LastEditors: joe 847304926@qq.com
# @LastEditTime: 2025-07-06 17:18:04
# @FilePath: \\wsl.localhost\Ubuntu-22.04\home\joe\wind_project\WindSimProj\backend\batch_2\atmospheric_run_20250627_230932\create_excel_summary_report.py
# @Description: 
# 
# Copyright (c) 2025 by joe, All Rights Reserved.

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成75m边界层Excel汇总报告（修复版本）
"""

import os
import json
import numpy as np
import pandas as pd
import argparse
import sys
from datetime import datetime
def set_font():
    import matplotlib.pyplot as plt
    import matplotlib.font_manager as fm
    chinese_pref = [
        'Noto Sans CJK SC',    # Ubuntu 推荐
        'WenQuanYi Micro Hei',
        'SimHei',              # Windows
        'Microsoft YaHei',
        'PingFang SC'          # macOS
    ]
    # 在已安装字体中就地取一个
    installed = {f.name for f in fm.fontManager.ttflist}
    for font in chinese_pref:
        if font in installed:
            plt.rcParams['font.sans-serif'] = [font]
            break
    else:
        # 兜底
        plt.rcParams['font.sans-serif'] = ['DejaVu Sans']
    plt.rcParams['axes.unicode_minus'] = False
def create_excel_summary_report(input_dir):
    """创建Excel汇总报告"""
    
    base_input_dir = os.path.abspath(input_dir)
    
    # 收集所有工况的数据
    all_data = []
    
    case_items = [item for item in os.listdir(base_input_dir) 
                  if os.path.isdir(os.path.join(base_input_dir, item)) and item.isdigit()]
    
    print(f"正在收集 {len(case_items)} 个工况的数据...")
    
    for item_name in sorted(case_items, key=int):
        item_path = os.path.join(base_input_dir, item_name)
        comparison_file = os.path.join(item_path, 'boundary_layer_75m', 'comparison_analysis.json')
        info_path = os.path.join(item_path, "info.json")
        
        if os.path.exists(comparison_file) and os.path.exists(info_path):
            try:
                with open(comparison_file, 'r', encoding='utf-8') as f:
                    comp_data = json.load(f)
                
                with open(info_path, 'r') as f:
                    info_data = json.load(f)
                
                wind_speed = info_data.get("wind", {}).get("speed", 10)
                wind_angle = info_data.get("wind", {}).get("angle", 0)
                
                our_results = comp_data['our_algorithm']['results']
                metodynwt_results = comp_data['metodynwt']
                field_results = comp_data['field_measurements']
                
                # 提取各地形的详细数据
                row_data = {
                    '工况编号': int(item_name),
                    '风速(m/s)': wind_speed,
                    '风向(°)': wind_angle,
                    '测量高度(m)': 75,
                }
                
                # 各地形的风速和误差
                terrain_names = ['山脊', '山谷', '平地']
                terrain_errors_our = []
                terrain_errors_met = []
                
                for terrain in terrain_names:
                    if terrain in our_results['measurement_points']:
                        # 实测值
                        field_speed = field_results[terrain]['speed']
                        # 自研算法
                        our_speed = our_results['measurement_points'][terrain]['speed']
                        # MetodynWT
                        met_speed = metodynwt_results[terrain]['speed']
                        
                        # 计算误差
                        our_error = abs(our_speed - field_speed) / field_speed * 100
                        met_error = abs(met_speed - field_speed) / field_speed * 100
                        
                        terrain_errors_our.append(our_error)
                        terrain_errors_met.append(met_error)
                        
                        # 添加到行数据
                        terrain_short = terrain.replace('平地', '平地')
                        row_data[f'{terrain_short}_实测(m/s)'] = round(field_speed, 3)
                        row_data[f'{terrain_short}_自研(m/s)'] = round(our_speed, 3)
                        row_data[f'{terrain_short}_WT(m/s)'] = round(met_speed, 3)
                        row_data[f'{terrain_short}_自研误差(%)'] = round(our_error, 2)
                        row_data[f'{terrain_short}_WT误差(%)'] = round(met_error, 2)
                
                # 平均误差
                our_avg_error = np.mean(terrain_errors_our) if terrain_errors_our else 0
                met_avg_error = np.mean(terrain_errors_met) if terrain_errors_met else 0
                
                # 计算时长
                our_time = comp_data['our_algorithm']['computation_time_hours']
                met_time = comp_data['metodynwt']['computation_time_hours']
                time_diff_percent = abs(our_time - met_time) / met_time * 100
                
                # 添加汇总数据
                row_data.update({
                    '自研平均误差(%)': round(our_avg_error, 2),
                    'WT平均误差(%)': round(met_avg_error, 2),
                    '自研计算时长(h)': round(our_time, 3),
                    'WT计算时长(h)': round(met_time, 3),
                    '时长差异(%)': round(time_diff_percent, 2),
                    '自研精度达标': '是' if our_avg_error <= 3 else '否',
                    'WT精度达标': '是' if met_avg_error <= 3 else '否',
                    '效率达标': '是' if time_diff_percent <= 5 else '否',
                    '精度优势': '自研' if our_avg_error < met_avg_error else 'WT' if met_avg_error < our_avg_error else '相当'
                })
                
                all_data.append(row_data)
                
            except Exception as e:
                print(f"处理工况 {item_name} 时出错: {e}")
    
    if not all_data:
        print("没有找到有效数据")
        return False
    
    # 创建DataFrame
    df = pd.DataFrame(all_data)
    
    # 创建Excel文件
    output_file = os.path.join(base_input_dir, f"75m边界层算法对比汇总报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    
    # 修复版本 - 使用多个写入操作而不是模式覆盖
    create_excel_workbook(df, output_file)
    
    print(f"Excel汇总报告已保存: {output_file}")
    
    # 格式化Excel
    format_excel_report(output_file)
    
    return True

def create_excel_workbook(df, output_file):
    """创建Excel工作簿（修复版本）"""
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # 1. 详细数据表
        df.to_excel(writer, sheet_name='详细对比数据', index=False)
        
        # 2. 统计摘要表
        create_summary_sheet_data(df, writer)
        
        # 3. 精度分析表 - 分为两个独立的表
        create_wind_speed_analysis(df, writer)
        create_wind_direction_analysis(df, writer)
        
        # 4. 效率分析表
        create_efficiency_analysis_sheet(df, writer)
        
        # 5. 地形分析表
        create_terrain_analysis_sheet(df, writer)

def create_summary_sheet_data(df, writer):
    """创建统计摘要表（修复版本）"""
    
    summary_data = {
        '项目': [
            '总工况数',
            '测量高度(m)',
            '自研算法平均误差(%)',
            'MetodynWT平均误差(%)',
            '自研算法平均时长(h)',
            'MetodynWT平均时长(h)',
            '平均时长差异(%)',
            '自研算法精度达标率(%)',
            'MetodynWT精度达标率(%)',
            '效率达标率(%)',
            '自研算法精度优势工况数',
            'MetodynWT精度优势工况数',
            '综合评价'
        ],
        '数值': [
            len(df),
            75,
            round(df['自研平均误差(%)'].mean(), 2),
            round(df['WT平均误差(%)'].mean(), 2),
            round(df['自研计算时长(h)'].mean(), 3),
            round(df['WT计算时长(h)'].mean(), 3),
            round(df['时长差异(%)'].mean(), 2),
            round(sum(df['自研精度达标'] == '是') / len(df) * 100, 1),
            round(sum(df['WT精度达标'] == '是') / len(df) * 100, 1),
            round(sum(df['效率达标'] == '是') / len(df) * 100, 1),
            sum(df['精度优势'] == '自研'),
            sum(df['精度优势'] == 'WT'),
            '自研算法优势' if df['自研平均误差(%)'].mean() < df['WT平均误差(%)'].mean() else 'MetodynWT优势'
        ]
    }
    
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='统计摘要', index=False)

def create_wind_speed_analysis(df, writer):
    """创建按风速的精度分析表"""
    
    # 按风速分组的精度分析
    wind_speed_groups = df.groupby('风速(m/s)').agg({
        '自研平均误差(%)': ['mean', 'std', 'min', 'max'],
        'WT平均误差(%)': ['mean', 'std', 'min', 'max']
    }).round(3)
    
    # 扁平化列名
    wind_speed_groups.columns = ['_'.join(col).strip() for col in wind_speed_groups.columns]
    wind_speed_groups = wind_speed_groups.reset_index()
    
    # 重命名列以便理解
    wind_speed_groups = wind_speed_groups.rename(columns={
        '自研平均误差(%)_mean': '自研平均误差(%)',
        '自研平均误差(%)_std': '自研误差标准差',
        '自研平均误差(%)_min': '自研误差最小值',
        '自研平均误差(%)_max': '自研误差最大值',
        'WT平均误差(%)_mean': 'WT平均误差(%)',
        'WT平均误差(%)_std': 'WT误差标准差',
        'WT平均误差(%)_min': 'WT误差最小值',
        'WT平均误差(%)_max': 'WT误差最大值'
    })
    
    wind_speed_groups.to_excel(writer, sheet_name='按风速精度分析', index=False)

def create_wind_direction_analysis(df, writer):
    """创建按风向的精度分析表"""
    
    # 按风向分组（可能需要分组处理）
    # 将风向分为8个主要方向
    def categorize_wind_direction(angle):
        if 0 <= angle < 45 or 315 <= angle < 360:
            return 'N (0-45°, 315-360°)'
        elif 45 <= angle < 90:
            return 'NE (45-90°)'
        elif 90 <= angle < 135:
            return 'E (90-135°)'
        elif 135 <= angle < 180:
            return 'SE (135-180°)'
        elif 180 <= angle < 225:
            return 'S (180-225°)'
        elif 225 <= angle < 270:
            return 'SW (225-270°)'
        elif 270 <= angle < 315:
            return 'W (270-315°)'
        else:
            return 'Other'
    
    df_wind_dir = df.copy()
    df_wind_dir['风向区间'] = df_wind_dir['风向(°)'].apply(categorize_wind_direction)
    
    wind_dir_groups = df_wind_dir.groupby('风向区间').agg({
        '自研平均误差(%)': ['mean', 'std', 'count'],
        'WT平均误差(%)': ['mean', 'std', 'count']
    }).round(3)
    
    wind_dir_groups.columns = ['_'.join(col).strip() for col in wind_dir_groups.columns]
    wind_dir_groups = wind_dir_groups.reset_index()
    
    wind_dir_groups.to_excel(writer, sheet_name='按风向精度分析', index=False)

def create_efficiency_analysis_sheet(df, writer):
    """创建效率分析表"""
    
    efficiency_data = {
        '工况编号': df['工况编号'],
        '风速(m/s)': df['风速(m/s)'],
        '风向(°)': df['风向(°)'],
        '自研计算时长(h)': df['自研计算时长(h)'],
        'WT计算时长(h)': df['WT计算时长(h)'],
        '时长差异(%)': df['时长差异(%)'],
        '效率达标': df['效率达标'],
        '时长比值': (df['自研计算时长(h)'] / df['WT计算时长(h)']).round(4)
    }
    
    efficiency_df = pd.DataFrame(efficiency_data)
    efficiency_df.to_excel(writer, sheet_name='效率分析', index=False)
def create_terrain_analysis_sheet(df, writer):
    """创建地形分析表（正确地形名称版本）"""
    
    # ✅ 使用正确的地形名称
    possible_terrains = ['平地', '山谷', '山脊']  # 这是实际的键名
    
    available_terrains = []
    terrain_display_names = []
    
    # 检查哪些地形列实际存在
    for terrain in possible_terrains:
        error_col = f'{terrain}_自研误差(%)'
        if error_col in df.columns:
            available_terrains.append(terrain)
            # 显示名称加上实际特征说明
            if terrain == '平地':
                terrain_display_names.append('平地(山脊特征)')
            elif terrain == '山谷':
                terrain_display_names.append('山谷(山谷特征)')
            elif terrain == '山脊':
                terrain_display_names.append('山脊(平地特征)')
            else:
                terrain_display_names.append(terrain)
    
    if not available_terrains:
        print("警告：没有找到地形误差数据列")
        # 创建空的地形分析表
        terrain_df = pd.DataFrame({
            '地形类型': ['无数据'],
            '自研平均误差(%)': [0],
            'WT平均误差(%)': [0],
            '说明': ['未找到地形数据']
        })
        terrain_df.to_excel(writer, sheet_name='地形分析', index=False)
        return
    
    # 各地形的平均误差
    terrain_summary = {
        '地形类型': terrain_display_names,
        '自研平均误差(%)': [],
        'WT平均误差(%)': [],
        '自研标准差': [],
        'WT标准差': [],
        '自研精度达标率(%)': [],
        'WT精度达标率(%)': []
    }
    
    for terrain in available_terrains:
        our_error_col = f'{terrain}_自研误差(%)'
        wt_error_col = f'{terrain}_WT误差(%)'
        
        terrain_summary['自研平均误差(%)'].append(df[our_error_col].mean())
        terrain_summary['WT平均误差(%)'].append(df[wt_error_col].mean())
        terrain_summary['自研标准差'].append(df[our_error_col].std())
        terrain_summary['WT标准差'].append(df[wt_error_col].std())
        terrain_summary['自研精度达标率(%)'].append(sum(df[our_error_col] <= 3) / len(df) * 100)
        terrain_summary['WT精度达标率(%)'].append(sum(df[wt_error_col] <= 3) / len(df) * 100)
    
    terrain_df = pd.DataFrame(terrain_summary)
    terrain_df = terrain_df.round(3)
    terrain_df.to_excel(writer, sheet_name='地形分析', index=False)

def format_excel_report(file_path):
    """格式化Excel报告"""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = load_workbook(file_path)
        
        # 定义样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # 格式化每个工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 设置标题行样式
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            
            # 设置数据行样式
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border
                    cell.alignment = center_alignment
            
            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_path)
        print("Excel格式化完成")
        
    except ImportError:
        print("警告: 缺少openpyxl库，无法进行Excel格式化")
    except Exception as e:
        print(f"Excel格式化失败: {e}")

def main():
    parser = argparse.ArgumentParser(description="生成75m边界层Excel汇总报告")
    parser.add_argument("input_dir", help="包含数字子文件夹的主数据文件夹路径")
    
    args = parser.parse_args()
    
    success = create_excel_summary_report(args.input_dir)
    
    if success:
        print("\n✅ Excel汇总报告生成完成！")
    else:
        print("\n❌ Excel汇总报告生成失败")
        sys.exit(1)

if __name__ == "__main__":
    main()